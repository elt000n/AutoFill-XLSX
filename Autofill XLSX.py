import calendar
import json
from datetime import datetime, date
import openpyxl
import os
import subprocess

# --- Загрузка конфига ---
CONFIG_FILE = "config.json"

try:
    with open(CONFIG_FILE, encoding="utf-8") as f:
        cfg = json.load(f)
    print(">>> Конфиг успешно загружен.")
except FileNotFoundError:
    print(f"❌ Не найден файл конфига '{CONFIG_FILE}'! Положите его в папку с скриптом.")
    input("Нажмите ENTER для выхода...")
    exit()
except json.JSONDecodeError as e:
    print(f"❌ Ошибка в конфиге (неверный JSON): {e}")
    input("Нажмите ENTER для выхода...")
    exit()

# --- Короткие алиасы чтобы не писать cfg["..."]["..."] везде ---
file_baza  = cfg["файлы"]["база"]
file_blank = cfg["файлы"]["бланк"]

B = cfg["база_строки"]
BL = cfg["бланк_строки"]
VAL = cfg["значения"]

ГРУППА_1 = set(cfg["дни_группа_1"])  # обычно пн/ср
ГРУППА_2 = set(cfg["дни_группа_2"])  # обычно вт/чт

# --- Текущая дата ---
now = datetime.now()
Текущий_год  = now.year
Текущий_месяц = now.month

MONTHS = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май",    6: "июнь",    7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
}

RU_MAP = {
    "mon": "пн", "tue": "вт", "wed": "ср", "thu": "чт",
    "fri": "пт", "sat": "сб", "sun": "вс"
}

# --- Безопасная запись ---
def safe_write(sheet, row, column, value):
    cell = sheet.cell(row=row, column=column)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            first_cell = sheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            )
            first_cell.value = value
            return
    cell.value = value

# --- Основной цикл ---
print(f"\n>>> База: '{file_baza}' | Бланк: '{file_blank}'")
print(f">>> Текущий месяц: {MONTHS[Текущий_месяц]} {Текущий_год}")

while True:
    print("\n" + "="*50)
    print("Ожидание ввода ученика")
    print("="*50)

    user_query = input("Введите фамилию ученика (или 'выход' для закрытия): ")
    if user_query.lower() in ["выход", "exit", "quit"]:
        print("Программа завершена.")
        break

    if not user_query.strip():
        print("❌ Вы ничего не ввели!")
        continue

    # --- Открываем базу ---
    try:
        baza_doc = openpyxl.load_workbook(file_baza, data_only=True)
    except FileNotFoundError:
        print(f"❌ Не найден файл базы: '{file_baza}'")
        continue

    # --- Ищем лист текущего месяца ---
    search_month = MONTHS[Текущий_месяц]
    target_sheet = next(
        (s for s in baza_doc.sheetnames if search_month.lower() in s.lower()),
        None
    )
    sheet_baza = baza_doc[target_sheet] if target_sheet else baza_doc.active

    if not target_sheet:
        print(f"⚠️ Лист '{search_month}' не найден, использую активный лист.")

    # --- Ищем строку ученика ---
    student_row = None
    for r in range(B["строка_учеников_начало"], B["строка_учеников_конец"]):
        cell_val = sheet_baza.cell(row=r, column=B["колонка_имя"]).value
        if cell_val and user_query.lower() in str(cell_val).lower():
            student_row = r
            full_student_name = cell_val
            break

    if not student_row:
        print(f"❌ Ученик '{user_query}' не найден в базе за {search_month}!")
        continue

    # --- Читаем посещения и определяем дни ученика ---
    attendance_data = {}
    day_of_week_map = {}
    student_days = set()

    col_start = B["колонки_посещений_начало"]
    col_end   = B["колонки_посещений_конец"]

    for col in range(col_start, col_end):
        dow      = sheet_baza.cell(row=B["строка_дней_недели"], column=col).value
        date_num = sheet_baza.cell(row=B["строка_дат"], column=col).value
        mark_val = sheet_baza.cell(row=student_row, column=col).value

        if date_num is None:
            continue
        try:
            date_num = int(date_num)
        except (ValueError, TypeError):
            continue

        dow_str = str(dow).strip().lower() if dow else ""
        day_of_week_map[date_num] = dow_str

        if mark_val is not None:
            try:
                attendance_data[date_num] = int(mark_val)
            except (ValueError, TypeError):
                attendance_data[date_num] = str(mark_val).strip()

            if dow_str in ГРУППА_1:
                student_days.update(ГРУППА_1)
            elif dow_str in ГРУППА_2:
                student_days.update(ГРУППА_2)

    if not student_days:
        student_days = set(
            day_of_week_map[d] for d in attendance_data if d in day_of_week_map
        )

    print(f">>> Ученик: {full_student_name} | Дни: {student_days} | Отметок: {len(attendance_data)}")

    # --- Строим список дат для бланка ---
    _, days_in_month = calendar.monthrange(Текущий_год, Текущий_месяц)
    student_dates = []

    for day in range(1, days_in_month + 1):
        weekday = date(Текущий_год, Текущий_месяц, day).strftime("%a").lower()
        ru_day  = RU_MAP.get(weekday, "")
        if ru_day in student_days:
            student_dates.append((day, ru_day))

    print(f">>> Дат для бланка: {len(student_dates)}")

    # --- Открываем бланк ---
    try:
        blank_doc = openpyxl.load_workbook(file_blank)
    except FileNotFoundError:
        print(f"❌ Не найден файл бланка: '{file_blank}'")
        continue

    sheet = blank_doc.worksheets[0]

    row_start = BL["строки_занятий_начало"]
    row_end   = BL["строки_занятий_конец"]

    # --- Очистка бланка ---
    for r in range(row_start, row_end):
        safe_write(sheet, r, BL["колонка_дата"],               None)
        safe_write(sheet, r, BL["колонка_проведено"],          None)
        safe_write(sheet, r, BL["колонка_отмена_родитель"],    None)
        safe_write(sheet, r, BL["колонка_отмена_специалист"],  None)

    # --- Заполняем даты и посещения ---
    for i, (day_num, _) in enumerate(student_dates):
        r = row_start + i
        if r >= row_end:
            break

        cell_date = date(Текущий_год, Текущий_месяц, day_num)
        safe_write(sheet, r, BL["колонка_дата"], cell_date)

        if day_num in attendance_data:
            val = attendance_data[day_num]
            if val == VAL["проведено"]:
                safe_write(sheet, r, BL["колонка_проведено"], VAL["проведено"])
            elif val == VAL["отмена_родителем"]:
                safe_write(sheet, r, BL["колонка_отмена_родитель"], VAL["отмена_родителем"])
            elif val == VAL["отмена_специалистом"]:
                safe_write(sheet, r, BL["колонка_отмена_специалист"], VAL["отмена_специалистом"])

    # --- Финансовый блок ---
    v_nach = sheet_baza.cell(row=student_row, column=B["колонка_остаток_начало"]).value
    v_opl  = sheet_baza.cell(row=student_row, column=B["колонка_оплата"]).value
    v_dat  = sheet_baza.cell(row=student_row, column=B["колонка_дата_оплаты"]).value

    sheet.cell(row=BL["строка_остаток_начало"], column=BL["колонка_остаток_начало"]).value = int(v_nach or 0)
    sheet.cell(row=BL["строка_оплата"],         column=BL["колонка_оплата"]).value         = int(v_opl  or 0)
    if v_dat is not None:
        sheet.cell(row=BL["строка_оплата"], column=BL["колонка_дата_оплаты"]).value = v_dat

    s_провед = sum(sheet.cell(row=r, column=BL["колонка_проведено"]).value or 0
                   for r in range(row_start, row_end))
    s_отмена = sum(sheet.cell(row=r, column=BL["колонка_отмена_родитель"]).value or 0
                   for r in range(row_start, row_end))

    itog = int(v_nach or 0) + int(v_opl or 0) - s_провед - s_отмена
    sheet.cell(row=BL["строка_итог"], column=BL["колонка_итог"]).value = itog

    print(f">>> Начало: {v_nach} | Оплата: {v_opl} | Проведено: {s_провед} | Отмены: {s_отмена} | Остаток: {itog}")

    # --- Сохранение ---
    while True:
        try:
            blank_doc.save(file_blank)
            blank_doc.close()
            print(f"✅ Готово! Данные {full_student_name} разнесены.")
            break
        except PermissionError:
            print("⚠️ Файл заблокирован! Закрываю Excel...")
            os.system("taskkill /f /im excel.exe >nul 2>&1")

    print("👉 Открываю Excel для скриншота...")
    try:
        subprocess.Popen(["start", "", file_blank], shell=True)
    except Exception:
        print("ℹ️ Не удалось открыть файл автоматически.")

    input("\nПосле скриншота нажмите ENTER...")

    print("🔄 Закрываю Excel, готовлюсь к следующему ученику...")
    os.system("taskkill /f /im excel.exe >nul 2>&1")